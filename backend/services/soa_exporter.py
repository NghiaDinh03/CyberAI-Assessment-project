"""Statement of Applicability (SoA) .xlsx exporter (Phase 3).

Reads the controls catalog (ISO 27001 or TCVN 11930) and assessment state from
``data/assessments/{assessment_id}.json`` to produce a standards-compliant
SoA spreadsheet.

Dependencies: ``openpyxl`` (already in requirements.txt for Phase 0).
"""

from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_CATEGORY_FONT = Font(name="Calibri", size=11, bold=True)
_BODY_FONT = Font(name="Calibri", size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Status mapping from numeric score to human-readable label (English for ISO)
_STATUS_MAP_EN = {
    0: "Not Implemented",
    1: "Initial / Ad Hoc",
    2: "Repeatable",
    3: "Defined",
    4: "Managed & Measurable",
    5: "Optimized",
}

# Status mapping for TCVN 11930 (Vietnamese)
_STATUS_MAP_VI = {
    0: "Chưa triển khai",
    1: "Ban đầu / Tự phát",
    2: "Lặp lại được",
    3: "Đã định nghĩa",
    4: "Được quản lý & Đo lường",
    5: "Tối ưu hóa",
}

# Column layout for ISO 27001 (English)
_COLUMNS_ISO = [
    ("Control ID", 14),
    ("Control Name", 40),
    ("Category", 20),
    ("Weight", 12),
    ("Applicable", 12),
    ("Justification for Inclusion/Exclusion", 40),
    ("Implementation Status", 22),
    ("Score (0-5)", 12),
    ("Evidence", 35),
    ("Notes", 30),
]

# Column layout for TCVN 11930 (Vietnamese)
_COLUMNS_TCVN = [
    ("Mã Kiểm soát", 14),
    ("Tên Biện pháp Kiểm soát", 40),
    ("Phân nhóm", 22),
    ("Trọng số", 12),
    ("Áp dụng", 12),
    ("Lý do Áp dụng / Loại trừ", 40),
    ("Trạng thái Triển khai", 22),
    ("Điểm số (0-5)", 12),
    ("Minh chứng", 35),
    ("Ghi chú khuyến nghị", 30),
]


def _flatten_controls(standard: str) -> List[dict]:
    """Flatten the nested category structure into a flat list with category info."""
    flat: List[dict] = []
    from services.controls_catalog import get_categories
    categories = get_categories(standard)
    for cat in categories:
        category_name = cat["category"]
        for ctrl in cat["controls"]:
            flat.append({
                "id": ctrl["id"],
                "label": ctrl["label"],
                "category": category_name,
                "weight": ctrl.get("weight", "medium"),
            })
    return flat


def _load_assessment(assessment_id: str) -> Optional[dict]:
    """Load assessment data from disk."""
    base = os.getenv("DATA_PATH", "./data")
    path = Path(base) / "assessments" / f"{assessment_id}.json"
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def _extract_control_scores(assessment: dict) -> Dict[str, dict]:
    """Extract per-control scoring data from an assessment record.

    Returns a dict keyed by control ID with:
        {score: int, status: str, evidence: list[str], notes: str}
    """
    scores: Dict[str, dict] = {}

    # Source 1: json_data from AI assessment result
    json_data = assessment.get("json_data") or {}
    controls_data = json_data.get("controls") or []
    for ctrl in controls_data:
        cid = ctrl.get("id") or ctrl.get("control_id", "")
        if cid:
            scores[cid] = {
                "score": ctrl.get("score", 0),
                "status": ctrl.get("status", ""),
                "notes": ctrl.get("recommendation", ctrl.get("notes", "")),
            }

    # Source 2: implemented_controls list (simpler fallback)
    sys_info = assessment.get("system_info") or {}
    compliance = sys_info.get("compliance") or {}
    implemented = compliance.get("implemented_controls") or []
    for cid in implemented:
        if cid not in scores:
            scores[cid] = {"score": 3, "status": "Implemented", "notes": ""}

    # Source 3: evidence_map
    evidence_map = sys_info.get("evidence_map") or assessment.get("evidence_map") or {}
    for cid, files in evidence_map.items():
        if cid in scores:
            scores[cid]["evidence"] = files
        else:
            scores[cid] = {"score": 0, "status": "", "notes": "", "evidence": files}

    return scores


def generate_soa_xlsx(
    assessment_id: Optional[str] = None,
    implemented_controls: Optional[List[str]] = None,
    org_name: str = "",
) -> bytes:
    """Generate a Statement of Applicability .xlsx file supporting dynamic standards.

    Args:
        assessment_id: If provided, loads scoring data and standard from the assessment.
        implemented_controls: Fallback list of implemented control IDs.
        org_name: Organization name for the header.

    Returns:
        Raw .xlsx bytes ready for HTTP streaming.
    """
    standard = "iso27001"
    control_scores: Dict[str, dict] = {}

    # Load assessment data if available
    if assessment_id:
        assessment = _load_assessment(assessment_id)
        if assessment:
            standard = assessment.get("standard", "iso27001")
            control_scores = _extract_control_scores(assessment)
            if not org_name:
                sys_info = assessment.get("system_info") or {}
                org_info = sys_info.get("organization") or {}
                org_name = org_info.get("name", "")

    # Fallback standard based on implemented controls prefixes if no ID
    elif implemented_controls:
        # TCVN controls start with NW., SV., APP., DAT., MNG.
        if any(c.split(".")[0] in ("NW", "SV", "APP", "DAT", "MNG") for c in implemented_controls):
            standard = "tcvn11930"
        for cid in implemented_controls:
            control_scores[cid] = {"score": 3, "status": "Implemented", "notes": ""}

    is_tcvn = standard == "tcvn11930"
    columns = _COLUMNS_TCVN if is_tcvn else _COLUMNS_ISO
    status_map = _STATUS_MAP_VI if is_tcvn else _STATUS_MAP_EN

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement of Applicability"

    # ── Title rows ───────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    if is_tcvn:
        title_cell.value = "Bản Tuyên bố Áp dụng — TCVN 11930:2017 (5 Cấp độ ATTT)"
    else:
        title_cell.value = "Statement of Applicability — ISO/IEC 27001:2022 Annex A"
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    meta_cell = ws["A2"]
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    org_label = "Tổ chức" if is_tcvn else "Organization"
    gen_label = "Xuất ngày" if is_tcvn else "Generated"
    meta_cell.value = f"{org_label}: {org_name or '—'}  |  {gen_label}: {generated}"
    meta_cell.font = Font(name="Calibri", size=10, italic=True)
    meta_cell.alignment = Alignment(horizontal="center")

    # ── Header row ───────────────────────────────────────────────────
    header_row = 4
    for col_idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ────────────────────────────────────────────────────
    controls = _flatten_controls(standard)
    current_row = header_row + 1
    last_category = ""

    for ctrl in controls:
        # Category separator row
        if ctrl["category"] != last_category:
            last_category = ctrl["category"]
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(columns),
            )
            cat_cell = ws.cell(row=current_row, column=1, value=last_category)
            cat_cell.fill = _CATEGORY_FILL
            cat_cell.font = _CATEGORY_FONT
            cat_cell.border = _THIN_BORDER
            current_row += 1

        cid = ctrl["id"]
        score_data = control_scores.get(cid, {})
        score = score_data.get("score", 0)
        is_applicable = cid in control_scores or not control_scores
        evidence_files = score_data.get("evidence", [])

        weight = ctrl["weight"]
        if is_tcvn:
            weight_label = "Chủ chốt" if weight == "critical" else "Cao" if weight == "high" else "Trung bình" if weight == "medium" else "Thấp"
            app_label = "Có" if is_applicable else "Không"
            just_label = "Bắt buộc theo TCVN" if is_applicable else "Không áp dụng — cần giải trình loại trừ"
        else:
            weight_label = weight.capitalize()
            app_label = "Yes" if is_applicable else "No"
            just_label = "Required per Annex A" if is_applicable else "Not applicable — justification required"

        row_data = [
            cid,
            ctrl["label"],
            ctrl["category"],
            weight_label,
            app_label,
            just_label,
            status_map.get(score, f"Điểm {score}" if is_tcvn else f"Score {score}"),
            score,
            ", ".join(evidence_files) if evidence_files else "",
            score_data.get("notes", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.alignment = _WRAP
            cell.border = _THIN_BORDER

        current_row += 1

    # ── Summary row ──────────────────────────────────────────────────
    current_row += 1
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=3,
    )
    sum_title = "Tổng kết" if is_tcvn else "Summary"
    summary_cell = ws.cell(row=current_row, column=1, value=sum_title)
    summary_cell.font = Font(name="Calibri", size=11, bold=True)

    total = len(controls)
    applicable = sum(1 for c in controls if c["id"] in control_scores) if control_scores else total
    implemented = sum(1 for c in controls if control_scores.get(c["id"], {}).get("score", 0) >= 3)

    t_label = f"Tổng số: {total}" if is_tcvn else f"Total: {total}"
    a_label = f"Áp dụng: {applicable}" if is_tcvn else f"Applicable: {applicable}"
    i_label = f"Đã đạt (≥3): {implemented}" if is_tcvn else f"Implemented (≥3): {implemented}"

    ws.cell(row=current_row, column=4, value=t_label).font = _BODY_FONT
    ws.cell(row=current_row, column=5, value=a_label).font = _BODY_FONT
    ws.cell(row=current_row, column=7, value=i_label).font = _BODY_FONT
    pct = round((implemented / applicable) * 100, 1) if applicable > 0 else 0
    ws.cell(row=current_row, column=8, value=f"{pct}%").font = Font(name="Calibri", size=11, bold=True)

    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"

    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{current_row - 1}"

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
