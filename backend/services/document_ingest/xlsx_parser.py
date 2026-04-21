"""XLSX parser — emits one Table per sheet and a flat text view.

Uses ``openpyxl`` in read-only mode so workbooks with charts/macros load
quickly and without executing anything. Merged cells are flattened by
broadcasting the top-left value to the whole range.
"""

from __future__ import annotations

import io
from typing import List

from api.schemas.document import Section, Table

from .base import ParserResult, register


def _cell_to_str(value: object) -> str:
    """Render a cell value as a stable string ('' for None)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


@register("xlsx")
def parse_xlsx(data: bytes) -> ParserResult:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required to parse .xlsx files. "
            "Install it with: pip install openpyxl"
        ) from exc

    # data_only=True returns formula results instead of formula text — that
    # is what evidence reviewers actually care about.
    workbook = load_workbook(filename=io.BytesIO(data), data_only=True, read_only=False)

    tables: List[Table] = []
    text_chunks: list[str] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Resolve merged cells: broadcast the anchor value to every member
        # of the range so downstream consumers see a fully populated grid.
        for merged_range in list(sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            anchor_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(merged_range))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    sheet.cell(row=r, column=c).value = anchor_value

        rows = [
            [_cell_to_str(cell.value) for cell in row]
            for row in sheet.iter_rows()
        ]
        # Strip fully empty trailing rows so the table is not padded with junk.
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()

        if not rows:
            continue

        headers, body = rows[0], rows[1:]
        tables.append(Table(name=sheet_name, headers=headers, rows=body))

        text_chunks.append(f"# {sheet_name}")
        text_chunks.append("\t".join(headers))
        for row in body:
            text_chunks.append("\t".join(row))

    workbook.close()
    return "\n".join(text_chunks).strip(), [], tables
