"""Unit tests for the SoA Exporter (Phase 3).

Tests verify:
1. Blank SoA generation (no assessment data)
2. SoA with implemented_controls list
3. SoA with assessment_id (mocked assessment file)
4. Valid .xlsx output (parseable by openpyxl)
5. HTTP endpoint returns correct content-type
"""

from __future__ import annotations

import json
import os
import pathlib
import sys

# Settings class is evaluated at import time — env vars must be set before
# any module that transitively imports core.config is loaded.
os.environ.setdefault("JWT_SECRET", "test-secret-at-least-32-characters-long!")
os.environ.setdefault("DEBUG", "true")

import pytest

BACKEND = pathlib.Path(__file__).resolve().parents[1]
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))


@pytest.fixture
def data_path(tmp_path, monkeypatch):
    monkeypatch.setenv("DATA_PATH", str(tmp_path))
    return tmp_path


def _create_assessment(data_path, assessment_id: str, **overrides) -> dict:
    """Write a minimal assessment JSON file for testing."""
    assessments_dir = data_path / "assessments"
    assessments_dir.mkdir(parents=True, exist_ok=True)

    assessment = {
        "id": assessment_id,
        "status": "completed",
        "system_info": {
            "organization": {"name": "Test Corp"},
            "compliance": {
                "implemented_controls": ["A.5.1", "A.5.2", "A.8.1"],
            },
            "evidence_map": {
                "A.5.1": ["policy.pdf", "review.docx"],
            },
        },
        "json_data": {
            "controls": [
                {"id": "A.5.1", "score": 4, "status": "Managed", "recommendation": "Continue monitoring"},
                {"id": "A.5.2", "score": 3, "status": "Defined", "recommendation": "Document roles"},
                {"id": "A.8.1", "score": 5, "status": "Optimized", "recommendation": ""},
            ],
        },
    }
    assessment.update(overrides)

    path = assessments_dir / f"{assessment_id}.json"
    path.write_text(json.dumps(assessment, ensure_ascii=False, indent=2), encoding="utf-8")
    return assessment


# ── Service-level tests ──────────────────────────────────────────────

class TestSoAExporterService:

    def test_blank_soa_generates_valid_xlsx(self, data_path):
        from services.soa_exporter import generate_soa_xlsx

        xlsx_bytes = generate_soa_xlsx()
        assert len(xlsx_bytes) > 0

        # Verify it's a valid xlsx
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.title == "Statement of Applicability"
        # Header row at row 4
        assert ws.cell(row=4, column=1).value == "Control ID"
        assert ws.cell(row=4, column=2).value == "Control Name"

    def test_soa_with_implemented_controls(self, data_path):
        from services.soa_exporter import generate_soa_xlsx

        xlsx_bytes = generate_soa_xlsx(
            implemented_controls=["A.5.1", "A.5.15"],
            org_name="Acme Corp",
        )

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        # Check org name in metadata row
        assert "Acme Corp" in (ws.cell(row=2, column=1).value or "")

        # Find A.5.1 row and verify it's marked as applicable
        found = False
        for row in ws.iter_rows(min_row=5, max_col=8, values_only=False):
            if row[0].value == "A.5.1":
                assert row[4].value == "Yes"  # Applicable
                assert row[7].value == 3  # Default score for implemented
                found = True
                break
        assert found, "A.5.1 not found in SoA output"

    def test_soa_with_assessment_id(self, data_path):
        from services.soa_exporter import generate_soa_xlsx

        _create_assessment(data_path, "test-assessment-001")

        xlsx_bytes = generate_soa_xlsx(assessment_id="test-assessment-001")

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        # Check org name pulled from assessment
        assert "Test Corp" in (ws.cell(row=2, column=1).value or "")

        # Find A.5.1 and verify score from assessment
        for row in ws.iter_rows(min_row=5, max_col=10, values_only=False):
            if row[0].value == "A.5.1":
                assert row[7].value == 4  # Score from json_data
                assert "policy.pdf" in (row[8].value or "")  # Evidence
                break

    def test_soa_has_all_93_controls(self, data_path):
        from services.soa_exporter import generate_soa_xlsx
        import re

        xlsx_bytes = generate_soa_xlsx()

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        # Count control rows — match "A.N.N" pattern (3-part IDs only),
        # skip category headers like "A.5 Tổ chức".
        ctrl_re = re.compile(r"^A\.\d+\.\d+$")
        control_ids = []
        for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
            val = row[0]
            if val and isinstance(val, str) and ctrl_re.match(val):
                control_ids.append(val)

        assert len(control_ids) == 93, f"Expected 93 controls, got {len(control_ids)}"

    def test_soa_nonexistent_assessment_returns_blank(self, data_path):
        from services.soa_exporter import generate_soa_xlsx

        # Should not raise, just return blank SoA
        xlsx_bytes = generate_soa_xlsx(assessment_id="nonexistent-id")
        assert len(xlsx_bytes) > 0


# ── Route-level tests ────────────────────────────────────────────────

# The iso27001 router transitively imports chromadb (via ChatService →
# VectorStore). Skip route-level tests when chromadb is not installed
# to keep the test suite runnable without heavy optional deps.
try:
    import chromadb  # noqa: F401
    _HAS_CHROMADB = True
except ImportError:
    _HAS_CHROMADB = False

_skip_no_chromadb = pytest.mark.skipif(
    not _HAS_CHROMADB,
    reason="chromadb not installed — iso27001 router cannot be imported",
)


@pytest.fixture
def iso_client(data_path):
    from fastapi import FastAPI
    from api.routes.iso27001 import router

    app = FastAPI()
    app.include_router(router, prefix="/api")
    from fastapi.testclient import TestClient
    return TestClient(app)


@_skip_no_chromadb
class TestSoAExportRoute:

    def test_export_blank_soa(self, iso_client):
        resp = iso_client.post("/api/iso27001/soa/export", json={})
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "SoA_ISO27001_" in resp.headers["content-disposition"]
        assert len(resp.content) > 0

    def test_export_with_controls(self, iso_client):
        resp = iso_client.post("/api/iso27001/soa/export", json={
            "implemented_controls": ["A.5.1", "A.5.2"],
            "org_name": "Test Org",
        })
        assert resp.status_code == 200
        assert len(resp.content) > 0

    def test_export_with_assessment(self, iso_client, data_path):
        _create_assessment(data_path, "route-test-001")
        resp = iso_client.post("/api/iso27001/soa/export", json={
            "assessment_id": "route-test-001",
        })
        assert resp.status_code == 200
        assert len(resp.content) > 0
